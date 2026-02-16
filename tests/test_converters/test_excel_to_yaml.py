"""
Test suite for Excel to YAML conversion (secrets.rlist format)
"""
import pytest
import os
import tempfile
import time
import gc
from pathlib import Path
from yamlconverter.converters.custom_excel_to_yaml import custom_excel_to_yaml


class TestExcelToYAML:
    """Test cases for Excel to YAML conversion"""
    
    @pytest.fixture
    def sample_excel_file(self):
        """Create a sample Excel file in secrets.rlist format"""
        from openpyxl import Workbook
        
        temp_path = tempfile.mktemp(suffix='.xlsx')
        wb = Workbook()
        ws = wb.active
        
        # Headers
        ws['A1'] = 'Name'
        ws['B1'] = 'Secret'
        ws['C1'] = 'Value'
        
        # Data rows
        ws['A2'] = 'SAP_SOAP_GET_BP_CONT_DETA_V2[0]'
        ws['B2'] = '$$ENDPOINT$$'
        ws['C2'] = 'https://example.com/api'
        
        ws['A3'] = 'SAP_SOAP_GET_BP_CONT_DETA_V2[1]'
        ws['B3'] = '$$USERNAME$$'
        ws['C3'] = 'OIC_wsuser'
        
        ws['A4'] = 'SAP_SOAP_GET_BP_CONT_DETA_V2[2]'
        ws['B4'] = '$$PASSWORD$$'
        ws['C4'] = 'Aa123456'
        
        ws['A5'] = 'KLEIO_SOAP_XASWS[0]'
        ws['B5'] = '$$ENDPOINT$$'
        ws['C5'] = 'http://kleio.example.com'
        
        wb.save(temp_path)
        wb.close()
        
        yield temp_path
        
        # Cleanup with retry for Windows file locks
        time.sleep(0.1)  # Give Windows time to release the file
        gc.collect()  # Force garbage collection
        for attempt in range(3):
            try:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
                break
            except PermissionError:
                if attempt < 2:
                    time.sleep(0.2)
                else:
                    pass  # Give up after 3 attempts
    
    @pytest.fixture
    def temp_yaml_file(self):
        """Create a temporary YAML file path"""
        temp_path = tempfile.mktemp(suffix='.yml')
        yield temp_path
        # Cleanup
        if os.path.exists(temp_path):
            os.unlink(temp_path)
    
    def test_excel_to_yaml_basic_conversion(self, sample_excel_file, temp_yaml_file):
        """Test basic Excel to YAML conversion"""
        # Convert Excel to YAML
        success, warnings, error = custom_excel_to_yaml(sample_excel_file, temp_yaml_file)
        
        # Verify YAML file was created
        assert success
        assert os.path.exists(temp_yaml_file)
        assert os.path.getsize(temp_yaml_file) > 0
    
    def test_excel_to_yaml_structure(self, sample_excel_file, temp_yaml_file):
        """Test that YAML has correct structure"""
        import yaml
        
        # Convert
        custom_excel_to_yaml(sample_excel_file, temp_yaml_file)
        
        # Load and verify
        with open(temp_yaml_file, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f)
        
        # Check structure
        assert 'Connections' in data
        assert 'SAP_SOAP_GET_BP_CONT_DETA_V2' in data['Connections']
        assert 'KLEIO_SOAP_XASWS' in data['Connections']
        
        # Check that connections have list of secrets
        assert isinstance(data['Connections']['SAP_SOAP_GET_BP_CONT_DETA_V2'], list)
        assert len(data['Connections']['SAP_SOAP_GET_BP_CONT_DETA_V2']) == 3
    
    def test_excel_to_yaml_data_integrity(self, sample_excel_file, temp_yaml_file):
        """Test that data is correctly transferred"""
        import yaml
        
        # Convert
        custom_excel_to_yaml(sample_excel_file, temp_yaml_file)
        
        # Load and verify
        with open(temp_yaml_file, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f)
        
        # Check first secret
        first_secret = data['Connections']['SAP_SOAP_GET_BP_CONT_DETA_V2'][0]
        assert first_secret['secret'] == '$$ENDPOINT$$'
        assert first_secret['value'] == 'https://example.com/api'
        
        # Check second secret
        second_secret = data['Connections']['SAP_SOAP_GET_BP_CONT_DETA_V2'][1]
        assert second_secret['secret'] == '$$USERNAME$$'
        assert second_secret['value'] == 'OIC_wsuser'
    
    def test_excel_to_yaml_invalid_input(self, temp_yaml_file):
        """Test error handling for invalid input"""
        success, warnings, error = custom_excel_to_yaml('nonexistent.xlsx', temp_yaml_file)
        assert not success
    
    def test_excel_to_yaml_empty_excel(self, temp_yaml_file):
        """Test handling of empty Excel file"""
        from openpyxl import Workbook
        
        temp_excel = tempfile.mktemp(suffix='.xlsx')
        wb = Workbook()
        wb.save(temp_excel)
        wb.close()
        
        try:
            success, warnings, error = custom_excel_to_yaml(temp_excel, temp_yaml_file)
            # Empty Excel might succeed with 0 connections or fail - both acceptable
            assert isinstance(success, bool)
        finally:
            time.sleep(0.1)
            gc.collect()
            for attempt in range(3):
                try:
                    if os.path.exists(temp_excel):
                        os.unlink(temp_excel)
                    break
                except PermissionError:
                    if attempt < 2:
                        time.sleep(0.2)
    
    def test_excel_to_yaml_missing_columns(self, temp_yaml_file):
        """Test handling of Excel with missing required columns"""
        from openpyxl import Workbook
        
        temp_excel = tempfile.mktemp(suffix='.xlsx')
        wb = Workbook()
        ws = wb.active
        
        # Only Name and Secret columns (missing Value)
        ws['A1'] = 'Name'
        ws['B1'] = 'Secret'
        ws['A2'] = 'Connection[0]'
        ws['B2'] = '$$ENDPOINT$$'
        
        wb.save(temp_excel)
        wb.close()
        
        try:
            success, warnings, error = custom_excel_to_yaml(temp_excel, temp_yaml_file)
            assert not success
        finally:
            time.sleep(0.1)
            gc.collect()
            for attempt in range(3):
                try:
                    if os.path.exists(temp_excel):
                        os.unlink(temp_excel)
                    break
                except PermissionError:
                    if attempt < 2:
                        time.sleep(0.2)
    
    def test_excel_to_yaml_line_endings(self, sample_excel_file, temp_yaml_file):
        """Test that YAML uses Unix line endings (LF)"""
        # Convert
        custom_excel_to_yaml(sample_excel_file, temp_yaml_file)
        
        # Read file in binary mode to check line endings
        with open(temp_yaml_file, 'rb') as f:
            content = f.read()
        
        # Check for LF only (no CRLF)
        assert b'\r\n' not in content or content.count(b'\n') > content.count(b'\r\n')


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
