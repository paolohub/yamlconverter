"""
Test suite for YAML to Excel conversion (secrets.rlist format)
"""
import pytest
import os
import tempfile
import time
import gc
from pathlib import Path
from yamlconverter.converters.custom_yaml_to_excel import custom_yaml_to_excel


class TestYAMLToExcel:
    """Test cases for YAML to Excel conversion"""
    
    @pytest.fixture
    def sample_yaml_content(self):
        """Sample YAML content in secrets.rlist format"""
        return """Connections:
  SAP_SOAP_GET_BP_CONT_DETA_V2:
    - secret: "$$ENDPOINT$$"
      value: "https://example.com/api"
    - secret: "$$USERNAME$$"
      value: "OIC_wsuser"
    - secret: "$$PASSWORD$$"
      value: "Aa123456"
  KLEIO_SOAP_XASWS:
    - secret: "$$ENDPOINT$$"
      value: "http://kleio.example.com"
    - secret: "$$USERNAME$$"
      value: "kleio_user"
"""
    
    @pytest.fixture
    def temp_yaml_file(self, sample_yaml_content):
        """Create a temporary YAML file"""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.yml', delete=False, encoding='utf-8') as f:
            f.write(sample_yaml_content)
            temp_path = f.name
        yield temp_path
        # Cleanup
        if os.path.exists(temp_path):
            os.unlink(temp_path)
    
    @pytest.fixture
    def temp_excel_file(self):
        """Create a temporary Excel file path"""
        temp_path = tempfile.mktemp(suffix='.xlsx')
        yield temp_path
        # Cleanup with retry for Windows file locks
        time.sleep(0.1)
        gc.collect()
        for attempt in range(3):
            try:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
                break
            except PermissionError:
                if attempt < 2:
                    time.sleep(0.2)
                else:
                    pass  # Give up after 3 attempts
    
    def test_yaml_to_excel_basic_conversion(self, temp_yaml_file, temp_excel_file):
        """Test basic YAML to Excel conversion"""
        # Convert YAML to Excel
        success, warnings, error = custom_yaml_to_excel(temp_yaml_file, temp_excel_file)
        
        # Verify Excel file was created
        assert success
        assert os.path.exists(temp_excel_file)
        assert os.path.getsize(temp_excel_file) > 0
    
    def test_yaml_to_excel_column_structure(self, temp_yaml_file, temp_excel_file):
        """Test that Excel file has correct column structure"""
        from openpyxl import load_workbook
        
        # Convert
        custom_yaml_to_excel(temp_yaml_file, temp_excel_file)
        
        # Load and verify
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        # Check headers
        assert ws['A1'].value == 'Name'
        assert ws['B1'].value == 'Secret'
        assert ws['C1'].value == 'Value'
        
        # Check that we have data rows
        assert ws.max_row > 1
        
        wb.close()
    
    def test_yaml_to_excel_data_integrity(self, temp_yaml_file, temp_excel_file):
        """Test that data is correctly transferred"""
        from openpyxl import load_workbook
        
        # Convert
        custom_yaml_to_excel(temp_yaml_file, temp_excel_file)
        
        # Load and verify
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        
        # Check first connection entry
        assert 'SAP_SOAP_GET_BP_CONT_DETA_V2[0]' in ws['A2'].value
        assert ws['B2'].value == '$$ENDPOINT$$'
        assert ws['C2'].value == 'https://example.com/api'
        
        wb.close()
    
    def test_yaml_to_excel_invalid_input(self, temp_excel_file):
        """Test error handling for invalid input"""
        success, warnings, error = custom_yaml_to_excel('nonexistent.yml', temp_excel_file)
        assert not success
    
    def test_yaml_to_excel_empty_file(self, temp_excel_file):
        """Test handling of empty YAML file"""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.yml', delete=False, encoding='utf-8') as f:
            f.write('')
            temp_yaml = f.name
        
        try:
            success, warnings, error = custom_yaml_to_excel(temp_yaml, temp_excel_file)
            assert not success
        finally:
            if os.path.exists(temp_yaml):
                os.unlink(temp_yaml)
    
    def test_yaml_to_excel_malformed_yaml(self, temp_excel_file):
        """Test handling of malformed YAML"""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.yml', delete=False, encoding='utf-8') as f:
            f.write('invalid: yaml: syntax:')
            temp_yaml = f.name
        
        try:
            success, warnings, error = custom_yaml_to_excel(temp_yaml, temp_excel_file)
            assert not success
        finally:
            if os.path.exists(temp_yaml):
                os.unlink(temp_yaml)


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
